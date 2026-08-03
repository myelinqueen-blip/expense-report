#!/usr/bin/env python3
"""
지출결의서 자동화 - Flask 백엔드 v7
실행: python server.py
접속: http://localhost:5000
"""

import os, io, json, copy
from datetime import date
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE  = os.path.join(BASE_DIR, '지출결의서_템플릿.xlsx')

# ─── 엑셀 생성 ────────────────────────────────────────────────────────────────

def build_excel(receipts: list, user_info: dict, label_mode: str,
                upload_date: date = None) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, GradientFill
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    if upload_date is None:
        upload_date = date.today()

    # 집행일: 다음달 5일
    if upload_date.month == 12:
        exec_year, exec_month = upload_date.year + 1, 1
    else:
        exec_year, exec_month = upload_date.year, upload_date.month + 1

    # 템플릿 로드 (원본 서식 완벽 유지)
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    # ── 날짜 헤더 입력 ──
    ws['C4'] = f"{upload_date.year}년"
    ws['D4'] = f"{upload_date.month}월"
    ws['E4'] = f"{upload_date.day}일"

    ws['C5'] = f"{upload_date.year}년"
    ws['D5'] = f"{upload_date.month}월"
    ws['E5'] = f"{upload_date.day}일"

    ws['C6'] = f"{exec_year}년"
    ws['D6'] = f"{exec_month}월"
    ws['E6'] = "5일"

    # ── 기안자 / 신청인 / 부서명 ──
    if user_info.get('writer'):
        ws['G4'] = user_info['writer']
    if user_info.get('requester'):
        ws['G5'] = user_info['requester']
    ws['G6'] = 'IMC사업본부'

    # ── 데이터 입력 (B9:J27) ──
    sorted_r = sorted(receipts, key=lambda r: r.get('date_sort', [0, 0, 0]))

    for i, r in enumerate(sorted_r[:19]):
        row = 9 + i
        ws.cell(row, 2).value  = r.get('date_str') or ''
        ws.cell(row, 3).value  = r.get('category', '')
        ws.cell(row, 4).value  = r.get('description', '')
        ws.cell(row, 8).value  = r.get('amount') or 0
        ws.cell(row, 10).value = i + 1   # 영수증 번호

    # ── Sheet2: 영수증 이미지 (A4 기준 6개/페이지, 3열×2행) ──
    if '영수증' in wb.sheetnames:
        del wb['영수증']
    iws = wb.create_sheet('영수증')

    # A4: 210×297mm. 3열 배치. 각 셀 너비 약 62mm
    # EMU: 1cm=360000, 1pt=12700
    # 열너비: openpyxl 단위 ≈ 문자수. 62mm ≈ 23 units
    COLS       = 3
    ROWS_PER_P = 2          # 페이지당 행 수
    PER_PAGE   = COLS * ROWS_PER_P  # 6개
    COL_W      = 23.0
    IMG_W_PX   = 170        # 이미지 최대 너비 px
    LBL_H_PT   = 20         # 레이블 행 높이 pt
    GAP_H_PT   = 8          # 페이지 구분 행 높이

    CIRCLES = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩',
               '⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲']

    # 열 너비 설정
    for ci in range(1, COLS + 1):
        iws.column_dimensions[get_column_letter(ci)].width = COL_W

    cur_excel_row = 1  # 현재 엑셀 행 (1-indexed)

    for idx, r in enumerate(sorted_r):
        page_pos = idx % PER_PAGE   # 페이지 내 위치 (0~5)
        col_pos  = page_pos % COLS  # 열 위치 (0~2)
        row_pos  = page_pos // COLS # 행 위치 (0~1)

        # 페이지 첫 번째 아이템이면 행 계산
        if page_pos == 0 and idx > 0:
            cur_excel_row += 1  # 페이지 구분 여백 행
            iws.row_dimensions[cur_excel_row].height = GAP_H_PT
            cur_excel_row += 1

        if col_pos == 0 and row_pos == 0:
            pass  # 페이지 시작, 아직 row 추가 안 함
        elif col_pos == 0 and row_pos > 0:
            pass  # 같은 페이지 2번째 행

        # 레이블 행 번호
        lbl_row = cur_excel_row + row_pos * 2
        img_row = lbl_row + 1

        # 첫 번째 열일 때만 행 높이 설정
        if col_pos == 0:
            iws.row_dimensions[lbl_row].height = LBL_H_PT

        # 레이블
        num = CIRCLES[idx] if idx < len(CIRCLES) else f'({idx+1})'
        if label_mode == 'with_desc':
            label_val = f"{num}  {r.get('date_str','')}  {r.get('description','')}"
        else:
            label_val = num

        lbl_cell = iws.cell(lbl_row, col_pos + 1)
        lbl_cell.value = label_val
        lbl_cell.font  = Font(name='맑은 고딕', size=9, bold=True)
        lbl_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 이미지 삽입 (원본 화질 유지)
        img_bytes = r.get('image_bytes', b'')
        if img_bytes:
            try:
                pil = PILImage.open(io.BytesIO(img_bytes))
                if pil.mode == 'RGBA':
                    bg = PILImage.new('RGB', pil.size, (255,255,255))
                    bg.paste(pil, mask=pil.split()[3]); pil = bg
                elif pil.mode not in ('RGB','L'):
                    pil = pil.convert('RGB')

                ow, oh = pil.size
                if ow > IMG_W_PX:
                    ratio = IMG_W_PX / ow
                    dw = IMG_W_PX; dh = int(oh * ratio)
                    pil = pil.resize((dw, dh), PILImage.LANCZOS)
                else:
                    dw, dh = ow, oh

                buf = io.BytesIO()
                pil.save(buf, format='PNG', optimize=False)
                buf.seek(0)

                xl_img = XLImage(buf)
                xl_img.width  = dw
                xl_img.height = dh
                iws.row_dimensions[img_row].height = dh * 0.75
                iws.add_image(xl_img, f'{get_column_letter(col_pos+1)}{img_row}')
            except Exception as e:
                iws.row_dimensions[img_row].height = 160
                iws.cell(img_row, col_pos+1).value = f'[이미지 오류]'
        else:
            iws.row_dimensions[img_row].height = 160

        # 페이지 내 마지막 아이템이면 cur_excel_row 업데이트
        if page_pos == PER_PAGE - 1 or idx == len(sorted_r) - 1:
            cur_excel_row += ROWS_PER_P * 2

    # A4 인쇄 설정
    iws.page_setup.paperSize   = 9
    iws.page_setup.orientation = 'portrait'
    iws.page_setup.fitToPage   = True
    iws.page_setup.fitToWidth  = 1
    iws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── 라우트 ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/generate', methods=['POST'])
def generate():
    try:
        meta       = json.loads(request.form.get('results', '[]'))
        user_info  = json.loads(request.form.get('user_info', '{}'))
        label_mode = request.form.get('label_mode', 'number_only')
    except Exception:
        return '파라미터 오류', 400

    files_map = {f.filename: f.read() for f in request.files.getlist('files')}
    for r in meta:
        r['image_bytes'] = files_map.get(r.get('filename',''), b'')

    try:
        xlsx = build_excel(meta, user_info, label_mode)
    except Exception:
        import traceback
        return f'Excel 오류:\n{traceback.format_exc()}', 500

    return send_file(
        io.BytesIO(xlsx),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='지출결의서_자동생성.xlsx'
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f'\n  지출결의서 자동화 v7  →  http://localhost:{port}\n')
    app.run(host='0.0.0.0', port=port, debug=False)
